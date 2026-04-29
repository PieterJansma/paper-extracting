from __future__ import annotations

from openpyxl import Workbook

from fix_molgenis_staging_types_callable import (
    _build_ref_index,
    _normalize_external_organisation_refs,
    coerce_ref,
)
from main_cohort import (
    _build_local_organisation_ref_map,
    _clear_organisation_parent_reference,
    _clear_self_organisation_reference,
    _ensure_organisation_ids,
    _registry_import_columns,
    _normalize_or_infer_resource_types,
    _resolve_local_organisation_ref,
    _resource_row_from_sections,
    _serialize_multi_value,
    _serialize_runtime_array_value,
)


def test_serialize_multi_value_flattens_and_cleans_bracketed_tokens() -> None:
    assert _serialize_multi_value('["Data access provider"]') == "Data access provider"
    assert _serialize_multi_value('Data access provider"]') == "Data access provider"


def test_runtime_ref_array_serializer_normalizes_and_filters_to_allowed_values() -> None:
    meta = {
        "column_type": "ref_array",
        "allowed_values": ["Data access provider", "Data controller"],
    }
    value = '["data access provider", "unknown role", "Data access provider\"]"]'
    assert _serialize_runtime_array_value("Agents", "role", meta, value) == "Data access provider"


def test_generated_organisation_ids_enable_local_reference_resolution() -> None:
    issues: list[dict[str, str]] = []
    organisations = [
        {
            "type": "organisation",
            "name": "UK Medical Research Council",
            "id": "",
        }
    ]
    _ensure_organisation_ids(
        organisations,
        paper_label="alspac",
        pdf_path="data/alspac.pdf",
        issues=issues,
    )
    assert organisations[0]["id"]
    assert issues

    ref_map = _build_local_organisation_ref_map(organisations)
    resolved = _resolve_local_organisation_ref("UK Medical Research Council", ref_map)
    assert resolved == organisations[0]["id"]


def test_self_organisation_reference_is_cleared_to_avoid_fk_error() -> None:
    row = {"id": "MRC", "organisation": "MRC"}
    _clear_self_organisation_reference(row)
    assert row["organisation"] == ""


def test_organisation_parent_reference_is_cleared_to_avoid_inherited_fk_error() -> None:
    row = {
        "id": "Department of Epidemiology",
        "type": "Organisation",
        "organisation": "UMCG",
        "other organisation": "",
    }
    _clear_organisation_parent_reference(row)
    assert row["organisation"] == ""
    assert row["other organisation"] == "UMCG"


def test_external_organisation_normalizer_does_not_backfill_blank_fk_from_name() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Organisations"
    ws.append(
        [
            "resource",
            "id",
            "type",
            "name",
            "organisation",
            "other organisation",
            "department",
            "website",
            "email",
            "logo",
            "role",
            "is lead organisation",
        ]
    )
    ws.append(
        [
            "10.1186/s12967-019-2122-x",
            "umcg-local",
            "Organisation",
            "UMCG",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    _normalize_external_organisation_refs(
        wb,
        {
            "exact": {"UMCG": "UMCG"},
            "norm": {"umcg": "UMCG"},
        },
    )

    assert ws.cell(row=2, column=5).value == ""


def test_organisation_refs_resolve_to_local_id_not_display_name() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Organisations"
    ws.append(
        [
            "resource",
            "id",
            "type",
            "name",
            "organisation",
            "other organisation",
            "department",
            "website",
            "email",
            "logo",
            "role",
            "is lead organisation",
        ]
    )
    ws.append(
        [
            "10.1186/s12967-019-2122-x",
            "umcg-local",
            "Organisation",
            "UMCG",
            "",
            "University Medical Center Groningen",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    ref_index = _build_ref_index(wb)

    assert coerce_ref("Resources", "publisher", "UMCG", ref_index) == "umcg-local"
    assert (
        coerce_ref(
            "Resources",
            "publisher",
            "University Medical Center Groningen",
            ref_index,
        )
        == "umcg-local"
    )


def test_resource_row_contract_has_non_empty_type() -> None:
    row, resource_ref = _resource_row_from_sections(
        "oncolifes",
        "data/oncolifes.pdf",
        {
            "task_overview": {
                "name": "OncoLifeS",
                "pid": "10.1186/s12967-019-2122-x",
                "type": [],
                "keywords": ["biobank"],
            },
            "task_design_structure": {},
            "task_population": {},
            "task_contributors": {},
            "task_areas_of_information": {},
            "task_linkage": {},
            "task_access_conditions": {},
            "task_information": {},
            "task_subpopulations": {"subpopulations": []},
            "task_collection_events": {"collection_events": []},
        },
        [],
    )

    assert resource_ref == "10.1186/s12967-019-2122-x"
    assert row["type"]
    assert _normalize_or_infer_resource_types({"name": "OncoLifeS biobank"})


def test_collections_import_columns_keep_inherited_id() -> None:
    registry = {
        "tables": {
            "Collections": {
                "extends": "Resources",
                "fields": {
                    "id": {"column_type": "string"},
                    "type": {"column_type": "ontology_array"},
                    "keywords": {"column_type": "string_array"},
                },
                "direct_field_names": ["type", "keywords"],
            }
        }
    }

    assert _registry_import_columns(registry, "Collections") == ["id", "type", "keywords"]
