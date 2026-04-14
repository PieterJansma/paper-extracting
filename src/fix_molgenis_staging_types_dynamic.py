from __future__ import annotations

import argparse
import csv
import difflib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

import fix_molgenis_staging_types_callable as legacy
from emx2_dynamic_runtime import build_runtime_registry
from llm_client import OpenAICompatibleClient


_CHOICE_INDEX_CACHE: Dict[str, Dict[str, Any]] = {}


@dataclass
class ChoiceMatchResult:
    raw: str
    mapped: Optional[str]
    method: str
    suggestions: List[str]


def _load_choice_index(source_path: str | None) -> Dict[str, Any]:
    if not source_path:
        return {}
    if source_path in _CHOICE_INDEX_CACHE:
        return _CHOICE_INDEX_CACHE[source_path]

    path = Path(source_path)
    if not path.is_file():
        return {}

    exact: Dict[str, str] = {}
    exact_multi: set[str] = set()
    norm: Dict[str, str] = {}
    norm_multi: set[str] = set()
    canonical_names: List[str] = []
    canonical_seen: set[str] = set()
    canonical_norm: Dict[str, str] = {}
    canonical_norm_multi: set[str] = set()

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            canonical = legacy._clean_string(row.get("name"))
            if not canonical:
                continue
            if canonical not in canonical_seen:
                canonical_seen.add(canonical)
                canonical_names.append(canonical)
            canonical_key = legacy._normalize_ref_token(canonical)
            if canonical_key:
                if canonical_key in canonical_norm and canonical_norm[canonical_key] != canonical:
                    canonical_norm_multi.add(canonical_key)
                else:
                    canonical_norm[canonical_key] = canonical
            tokens = {
                canonical,
                legacy._clean_string(row.get("label")),
                legacy._clean_string(row.get("acronym")),
                legacy._clean_string(row.get("code")),
                legacy._clean_string(row.get("ontologyTermURI")),
                legacy._clean_string(row.get("id")),
            }
            tokens.discard("")

            for token in tokens:
                if token in exact and exact[token] != canonical:
                    exact_multi.add(token)
                else:
                    exact[token] = canonical

                normalized = legacy._normalize_ref_token(token)
                if not normalized:
                    continue
                if normalized in norm and norm[normalized] != canonical:
                    norm_multi.add(normalized)
                else:
                    norm[normalized] = canonical

    for token in exact_multi:
        exact.pop(token, None)
    for token in norm_multi:
        norm.pop(token, None)
    for token in canonical_norm_multi:
        canonical_norm.pop(token, None)

    index = {
        "exact": exact,
        "norm": norm,
        "canonical": canonical_names,
        "canonical_norm": canonical_norm,
    }
    _CHOICE_INDEX_CACHE[source_path] = index
    return index


def _match_choice(
    source_path: str | None,
    raw_value: Any,
    *,
    fuzzy_cutoff: float = 0.96,
    fuzzy_margin: float = 0.02,
) -> ChoiceMatchResult:
    raw = legacy._clean_string(raw_value)
    if not raw or not source_path:
        return ChoiceMatchResult(raw=raw, mapped=None, method="empty", suggestions=[])

    choice_index = _load_choice_index(source_path)
    if not choice_index:
        return ChoiceMatchResult(raw=raw, mapped=None, method="no_index", suggestions=[])

    exact = choice_index.get("exact", {})
    if raw in exact:
        return ChoiceMatchResult(raw=raw, mapped=exact[raw], method="exact", suggestions=[])

    normalized = legacy._normalize_ref_token(raw)
    norm = choice_index.get("norm", {})
    if normalized and normalized in norm:
        return ChoiceMatchResult(raw=raw, mapped=norm[normalized], method="normalized", suggestions=[])

    if not normalized:
        return ChoiceMatchResult(raw=raw, mapped=None, method="empty_normalized", suggestions=[])

    canonical_norm = choice_index.get("canonical_norm", {})
    if normalized in canonical_norm:
        return ChoiceMatchResult(
            raw=raw,
            mapped=canonical_norm[normalized],
            method="normalized_canonical",
            suggestions=[],
        )

    score_pairs = [
        (token, difflib.SequenceMatcher(None, normalized, token).ratio())
        for token in canonical_norm.keys()
    ]
    score_pairs.sort(key=lambda item: item[1], reverse=True)
    best_token, best_score = score_pairs[0] if score_pairs else ("", 0.0)
    second_score = score_pairs[1][1] if len(score_pairs) > 1 else 0.0
    if best_score >= fuzzy_cutoff and (best_score - second_score) >= fuzzy_margin:
        return ChoiceMatchResult(
            raw=raw,
            mapped=canonical_norm.get(best_token),
            method="fuzzy_auto",
            suggestions=[],
        )

    suggestions: List[str] = []
    seen: set[str] = set()
    for token, _ in score_pairs[:5]:
        candidate = canonical_norm.get(token, "")
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        suggestions.append(candidate)
    return ChoiceMatchResult(raw=raw, mapped=None, method="unmapped", suggestions=suggestions)


def _resolve_choice_with_llm(
    *,
    raw: str,
    candidates: List[str],
    client: OpenAICompatibleClient,
    field_label: str,
) -> Optional[str]:
    if not candidates:
        return None
    system = (
        "You map extracted metadata values to one canonical schema value. "
        "Return ONLY JSON: {\"choice\": <candidate or null>}."
    )
    user = (
        f"Field: {field_label}\n"
        f"Raw value: {raw}\n"
        f"Candidates: {json.dumps(candidates, ensure_ascii=False)}\n"
        "Choose exactly one candidate if clearly matching, otherwise null."
    )
    try:
        text = client.chat(
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0.0,
            max_tokens=256,
            response_format={"type": "json_object"},
            max_retries=2,
        )
        payload = json.loads(text.strip())
        choice = payload.get("choice")
        if isinstance(choice, str) and choice in candidates:
            return choice
    except Exception:
        return None
    return None


def _map_choice(
    source_path: str | None,
    raw_value: Any,
    *,
    field_label: str,
    llm_client: Optional[OpenAICompatibleClient] = None,
    llm_choice_threshold: int = 16,
    llm_max_candidates: int = 5,
    llm_max_lookups: int = 50,
    llm_state: Optional[Dict[str, int]] = None,
) -> str:
    match = _match_choice(source_path, raw_value)
    if match.mapped:
        return match.mapped

    if not llm_client or not match.suggestions:
        return ""

    choice_index = _load_choice_index(source_path)
    choice_count = len(choice_index.get("canonical", []) or [])
    if choice_count <= llm_choice_threshold:
        return ""

    if llm_state is not None:
        used = int(llm_state.get("lookups", 0))
        if used >= llm_max_lookups:
            return ""
        llm_state["lookups"] = used + 1

    llm_choice = _resolve_choice_with_llm(
        raw=match.raw,
        candidates=match.suggestions[:llm_max_candidates],
        client=llm_client,
        field_label=field_label,
    )
    return llm_choice or ""


def _coerce_dynamic_ontology(
    meta: Dict[str, Any],
    value: Any,
    *,
    llm_client: Optional[OpenAICompatibleClient] = None,
    llm_choice_threshold: int = 16,
    llm_max_candidates: int = 5,
    llm_max_lookups: int = 50,
    llm_state: Optional[Dict[str, int]] = None,
) -> Any:
    mapped = _map_choice(
        meta.get("source_path"),
        legacy._extract_ontology_scalar(value),
        field_label=f"{meta['table_name']}.{meta['column_name']}",
        llm_client=llm_client,
        llm_choice_threshold=llm_choice_threshold,
        llm_max_candidates=llm_max_candidates,
        llm_max_lookups=llm_max_lookups,
        llm_state=llm_state,
    )
    if mapped:
        return mapped
    return legacy.coerce_ontology(meta["table_name"], meta["column_name"], value)


def _coerce_dynamic_ontology_array(
    meta: Dict[str, Any],
    value: Any,
    *,
    llm_client: Optional[OpenAICompatibleClient] = None,
    llm_choice_threshold: int = 16,
    llm_max_candidates: int = 5,
    llm_max_lookups: int = 50,
    llm_state: Optional[Dict[str, int]] = None,
) -> Any:
    raw_items = legacy._parse_array_items(value)
    if raw_items is None:
        return value

    items: list[str] = []
    seen: set[str] = set()
    for raw in raw_items:
        scalar = legacy._extract_ontology_scalar(raw)
        if not scalar:
            continue
        mapped = _map_choice(
            meta.get("source_path"),
            scalar,
            field_label=f"{meta['table_name']}.{meta['column_name']}",
            llm_client=llm_client,
            llm_choice_threshold=llm_choice_threshold,
            llm_max_candidates=llm_max_candidates,
            llm_max_lookups=llm_max_lookups,
            llm_state=llm_state,
        ) or scalar
        if not mapped or mapped in seen:
            continue
        seen.add(mapped)
        items.append(mapped)

    allowed = set(meta.get("allowed_values") or [])
    if allowed:
        items = [item for item in items if item in allowed]
    elif not meta.get("source_path"):
        return legacy.coerce_array(meta["table_name"], meta["column_name"], value)

    return ",".join(items)


def _coerce_dynamic_external_ref(meta: Dict[str, Any], value: Any) -> Any:
    scalar = legacy._extract_ref_scalar(value)
    if not scalar:
        return ""
    mapped = _map_choice(
        meta.get("source_path"),
        scalar,
        field_label=f"{meta['table_name']}.{meta['column_name']}",
    )
    return mapped or ""


def _coerce_dynamic_external_ref_array(meta: Dict[str, Any], value: Any) -> Any:
    raw_items = legacy._parse_array_items(value)
    if raw_items is None:
        return value

    items: list[str] = []
    seen: set[str] = set()
    for raw in raw_items:
        scalar = legacy._extract_ref_scalar(raw)
        if not scalar:
            continue
        mapped = _map_choice(
            meta.get("source_path"),
            scalar,
            field_label=f"{meta['table_name']}.{meta['column_name']}",
        )
        if not mapped or mapped in seen:
            continue
        seen.add(mapped)
        items.append(mapped)
    return ",".join(items)


def _build_schema_from_registry(registry: Dict[str, Any]) -> Dict[str, Dict[str, Dict[str, Any]]]:
    schema: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for table_name, table_meta in registry.get("tables", {}).items():
        fields = table_meta.get("fields", {})
        schema[table_name] = {}
        for column_name, meta in fields.items():
            enriched = dict(meta)
            enriched["table_name"] = table_name
            enriched["column_name"] = column_name
            schema[table_name][column_name] = enriched
    return schema


def _resolve_organisations_source_path(registry: Dict[str, Any]) -> str:
    for table_name in ("Organisations", "Agents"):
        field_meta = (
            registry.get("tables", {})
            .get(table_name, {})
            .get("fields", {})
            .get("organisation")
        )
        if field_meta and field_meta.get("source_path"):
            return str(field_meta["source_path"])
    return ""


def _normalize_non_ref_value(
    meta: Dict[str, Any],
    value: Any,
    *,
    llm_client: Optional[OpenAICompatibleClient] = None,
    llm_choice_threshold: int = 16,
    llm_max_candidates: int = 5,
    llm_max_lookups: int = 50,
    llm_state: Optional[Dict[str, int]] = None,
) -> Any:
    column_type = meta.get("column_type")
    if column_type == "ontology":
        return _coerce_dynamic_ontology(
            meta,
            value,
            llm_client=llm_client,
            llm_choice_threshold=llm_choice_threshold,
            llm_max_candidates=llm_max_candidates,
            llm_max_lookups=llm_max_lookups,
            llm_state=llm_state,
        )
    if column_type == "ontology_array":
        return _coerce_dynamic_ontology_array(
            meta,
            value,
            llm_client=llm_client,
            llm_choice_threshold=llm_choice_threshold,
            llm_max_candidates=llm_max_candidates,
            llm_max_lookups=llm_max_lookups,
            llm_state=llm_state,
        )
    if column_type == "string_array":
        return legacy.coerce_array(meta["table_name"], meta["column_name"], value)
    if column_type == "heading":
        return legacy.coerce_heading(value)
    if column_type == "date":
        return legacy.coerce_date(value)
    if column_type == "datetime":
        return legacy.coerce_datetime(value)
    if column_type == "email":
        return legacy.coerce_email(value)
    if column_type == "file":
        return legacy.coerce_file(value)
    if column_type == "int":
        return legacy.coerce_int(value, non_negative=False)
    if column_type == "non_negative_int":
        return legacy.coerce_int(value, non_negative=True)
    if column_type == "hyperlink":
        return legacy.coerce_hyperlink(meta["table_name"], meta["column_name"], value)
    if column_type == "bool":
        return legacy.coerce_bool(value)
    if column_type == "text":
        return legacy.coerce_passthrough_text(value)
    if column_type in legacy.PASSTHROUGH_TYPES:
        return legacy.coerce_passthrough(value)
    return value


def fix_workbook_dynamic(
    input_path: str | Path,
    *,
    output_path: str | Path | None = None,
    registry: Optional[Dict[str, Any]] = None,
    profile: str = "UMCGCohortsStaging",
    local_root: str | None = None,
    fallback_schema_csv: str | None = None,
    cache_dir: str | None = None,
    llm_client: Optional[OpenAICompatibleClient] = None,
    llm_choice_threshold: int = 16,
    llm_max_candidates: int = 5,
    llm_max_lookups: int = 50,
) -> Path:
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else input_path.with_name(input_path.stem + "_dynamic.xlsx")

    runtime_registry = registry or build_runtime_registry(
        profile,
        local_root=local_root,
        fallback_schema_csv=fallback_schema_csv,
        cache_dir=cache_dir,
    )
    schema = _build_schema_from_registry(runtime_registry)
    wb = load_workbook(input_path)
    llm_state = {"lookups": 0}

    organisations_source_path = _resolve_organisations_source_path(runtime_registry)
    if organisations_source_path:
        legacy._normalize_external_organisation_refs(
            wb,
            legacy._load_external_organisations_index(organisations_source_path),
        )

    for sheet_name in wb.sheetnames:
        if sheet_name not in schema:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            column_name = str(header)
            meta = schema[sheet_name].get(column_name)
            if not meta:
                continue
            column_type = meta.get("column_type")
            if column_type in {"ref", "refback"}:
                continue
            if column_type == "ref_array" and not meta.get("ref_schema"):
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = _normalize_non_ref_value(
                    meta,
                    cell.value,
                    llm_client=llm_client,
                    llm_choice_threshold=llm_choice_threshold,
                    llm_max_candidates=llm_max_candidates,
                    llm_max_lookups=llm_max_lookups,
                    llm_state=llm_state,
                )

    ref_index = legacy._build_ref_index(wb)
    for sheet_name in wb.sheetnames:
        if sheet_name not in schema:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            column_name = str(header)
            meta = schema[sheet_name].get(column_name)
            if not meta:
                continue
            column_type = meta.get("column_type")
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if column_type == "refback":
                    cell.value = ""
                elif column_type == "ref":
                    if meta.get("ref_schema") and meta.get("source_path"):
                        cell.value = _coerce_dynamic_external_ref(meta, cell.value)
                    else:
                        cell.value = legacy.coerce_ref(sheet_name, column_name, cell.value, ref_index)
                elif column_type == "ref_array":
                    if meta.get("ref_schema") and meta.get("source_path"):
                        cell.value = _coerce_dynamic_external_ref_array(meta, cell.value)
                    else:
                        cell.value = legacy.coerce_ref_array(sheet_name, column_name, cell.value, ref_index)

    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize workbook using schema-driven EMX2 runtime metadata.")
    parser.add_argument("input", help="Input workbook to fix")
    parser.add_argument("-o", "--output", default=None, help="Output workbook path")
    parser.add_argument("--profile", default="UMCGCohortsStaging", help="Profile name")
    parser.add_argument("--local-root", default=None, help="Optional local molgenis-emx2 checkout root")
    parser.add_argument("--fallback-schema-csv", default=None, help="Fallback schema CSV path")
    parser.add_argument("--cache-dir", default=None, help="EMX2 cache directory with fetched CSV files")
    args = parser.parse_args()

    output_path = fix_workbook_dynamic(
        args.input,
        output_path=args.output,
        profile=args.profile,
        local_root=args.local_root,
        fallback_schema_csv=args.fallback_schema_csv,
        cache_dir=args.cache_dir,
    )
    print(output_path)


if __name__ == "__main__":
    main()
