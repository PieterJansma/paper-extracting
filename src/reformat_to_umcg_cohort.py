from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Expected cohort structure
# -----------------------------

INHERITED_TABLES = {
    "Organisations": "Agents",
}

def load_model_structure(model_path: Path) -> Dict[str, List[str]]:
    model_df = pd.read_excel(model_path, dtype="object")
    required = {"tableName", "columnName"}
    missing = required - set(model_df.columns)
    if missing:
        raise SystemExit(f"Model file misses required columns: {sorted(missing)}")

    tables: Dict[str, List[str]] = {}
    for table_name in model_df["tableName"].dropna().unique().tolist():
        cols = model_df.loc[model_df["tableName"] == table_name, "columnName"].dropna().tolist()
        tables[str(table_name)] = [str(c) for c in cols]

    for child_table, parent_table in INHERITED_TABLES.items():
        if child_table not in tables or parent_table not in tables:
            continue
        merged = list(tables[parent_table])
        for col in tables[child_table]:
            if col not in merged:
                merged.append(col)
        tables[child_table] = merged
    return tables


# -----------------------------
# Source -> target mapping
# -----------------------------

SHEET_MAP = {
    "resources": "Resources",
    "subpopulations": "Subpopulations",
    "collection_events": "Collection events",
    "datasets": "Datasets",
    "organisations": "Organisations",
    "people": "Contacts",
    "publications": "Publications",
    "documentation": "Documentation",
}

# Exact column mapping from current extractor output to target cohort columns.
COLUMN_MAP = {
    "Resources": {
        "paper": "name",  # fallback label when better resource key is unavailable
        "task_overview.overview": "overview",
        "task_overview.hricore": "hricore",
        "task_overview.id": "id",
        "task_overview.pid": "pid",
        "task_overview.name": "name",
        "task_overview.acronym": "acronym",
        "task_overview.type": "type",
        "task_overview.cohort_type": "cohort type",
        "task_overview.website": "website",
        "task_overview.description": "description",
        "task_overview.keywords": "keywords",
        "task_overview.internal_identifiers": "internal identifiers",
        "task_overview.external_identifiers": "external identifiers",
        "task_overview.start_year": "start year",
        "task_overview.end_year": "end year",
        "task_overview.contact_email": "contact email",
        "task_overview.logo": "logo",
        "task_overview.issued": "issued",
        "task_overview.modified": "modified",
        "task_design_structure.design_and_structure": "design and structure",
        "task_design_structure.design": "design",
        "task_design_structure.design_description": "design description",
        "task_design_structure.design_schematic": "design schematic",
        "task_design_structure.data_collection_type": "data collection type",
        "task_design_structure.subpopulations": "subpopulations",
        "task_design_structure.collection_events": "collection events",
        "task_population.population": "population",
        "task_population.number_of_participants": "number of participants",
        "task_population.number_of_participants_with_samples": "number of participants with samples",
        "task_population.countries": "countries",
        "task_population.regions": "regions",
        "task_population.population_age_groups": "population age groups",
        "task_population.age_min": "age min",
        "task_population.age_max": "age max",
        "task_population.inclusion_criteria": "inclusion criteria",
        "task_population.other_inclusion_criteria": "other inclusion criteria",
        "task_population.exclusion_criteria": "exclusion criteria",
        "task_population.other_exclusion_criteria": "other exclusion criteria",
        "task_linkage.linkage": "linkage",
        "task_linkage.linkage_options": "linkage options",
        "task_access_conditions.informed_consent_type": "informed consent type",
        "task_access_conditions.access_rights": "access rights",
        "task_access_conditions.data_access_conditions": "data access conditions",
        "task_access_conditions.data_use_conditions": "data use conditions",
        "task_access_conditions.data_access_conditions_description": "data access conditions description",
        "task_access_conditions.data_access_fee": "data access fee",
        "task_access_conditions.release_type": "release type",
        "task_access_conditions.release_description": "release description",
        "task_information.funding_statement": "funding statement",
        "task_information.acknowledgements": "acknowledgements",
        "task_information.provenance_statement": "provenance statement",
        "task_information.theme": "theme",
        "task_information.applicable_legislation": "applicable legislation",
    },
    "Subpopulations": {
        "paper": "resource",
        "subpopulation.name": "name",
        "subpopulation.pid": "pid",
        "subpopulation.description": "description",
        "subpopulation.keywords": "keywords",
        "subpopulation.number_of_participants": "number of participants",
        "subpopulation.inclusion_start": "inclusion start",
        "subpopulation.inclusion_end": "inclusion end",
        "subpopulation.age_groups": "age groups",
        "subpopulation.age_min": "age min",
        "subpopulation.age_max": "age max",
        "subpopulation.main_medical_condition": "main medical condition",
        "subpopulation.comorbidity": "comorbidity",
        "subpopulation.countries": "countries",
        "subpopulation.regions": "regions",
        "subpopulation.inclusion_criteria": "inclusion criteria",
        "subpopulation.other_inclusion_criteria": "other inclusion criteria",
        "subpopulation.exclusion_criteria": "exclusion criteria",
        "subpopulation.other_exclusion_criteria": "other exclusion criteria",
        "subpopulation.issued": "issued",
        "subpopulation.modified": "modified",
        "subpopulation.theme": "theme",
        "subpopulation.access_rights": "access rights",
        "subpopulation.applicable_legislation": "applicable legislation",
    },
    "Collection events": {
        "paper": "resource",
        "collection_event.name": "name",
        "collection_event.pid": "pid",
        "collection_event.description": "description",
        "collection_event.subpopulations": "subpopulations",
        "collection_event.keywords": "keywords",
        "collection_event.start_date": "start date",
        "collection_event.end_date": "end date",
        "collection_event.age_groups": "age groups",
        "collection_event.number_of_participants": "number of participants",
        "collection_event.areas_of_information": "areas of information",
        "collection_event.data_categories": "data categories",
        "collection_event.sample_categories": "sample categories",
        "collection_event.standardized_tools": "standardized tools",
        "collection_event.standardized_tools_other": "standardized tools other",
        "collection_event.core_variables": "core variables",
        "collection_event.issued": "issued",
        "collection_event.modified": "modified",
        "collection_event.theme": "theme",
        "collection_event.access_rights": "access rights",
        "collection_event.applicable_legislation": "applicable legislation",
    },
    "Datasets": {
        "paper": "resource",
        "dataset.name": "name",
        "dataset.label": "label",
        "dataset.dataset_type": "dataset type",
        "dataset.unit_of_observation": "unit of observation",
        "dataset.keywords": "keywords",
        "dataset.description": "description",
        "dataset.number_of_rows": "number of rows",
        "dataset.since_version": "since version",
        "dataset.until_version": "until version",
    },
    "Agents": {
        "paper": "resource",
        "organisation.id": "id",
        "organisation.type": "type",
        "organisation.name": "name",
        "organisation.organisation": "organisation",
        "organisation.other_organisation": "other organisation",
        "organisation.department": "department",
        "organisation.website": "website",
        "organisation.email": "email",
        "organisation.logo": "logo",
        "organisation.role": "role",
    },
    "Organisations": {
        "paper": "resource",
        "organisation.id": "id",
        "organisation.type": "type",
        "organisation.name": "name",
        "organisation.organisation": "organisation",
        "organisation.other_organisation": "other organisation",
        "organisation.department": "department",
        "organisation.website": "website",
        "organisation.email": "email",
        "organisation.logo": "logo",
        "organisation.role": "role",
        "organisation.is_lead_organisation": "is lead organisation",
    },
    "Contacts": {
        "paper": "resource",
        "person.role": "role",
        "person.first_name": "first name",
        "person.last_name": "last name",
        "person.prefix": "prefix",
        "person.initials": "initials",
        "person.title": "title",
        "person.organisation": "organisation",
        "person.email": "email",
        "person.orcid": "orcid",
        "person.homepage": "homepage",
        "person.photo": "photo",
        "person.expertise": "expertise",
    },
    "Publications": {
        "paper": "resource",
        "publication.doi": "doi",
        "publication.title": "title",
        "publication.is_design_publication": "is design publication",
    },
    "Documentation": {
        "paper": "resource",
        "documentation.name": "name",
        "documentation.type": "type",
        "documentation.description": "description",
        "documentation.url": "url",
        "documentation.file": "file",
    },
}

# Target sheets that are not present in the current extractor output at all.
KNOWN_NOT_EXTRACTED = [
    "Variable values",
    "Profiles",
    "External identifiers",
    "Version",
    "Variable mappings",
    "Variables",
    "Internal identifiers",
    "Subpopulation counts",
]

# Target columns that cannot be reconstructed from current output with confidence.
KNOWN_PARTIAL_COLUMNS = {
    "Resources": [
        "organisations involved",
        "people involved",
        "contact point",
        "publications",
        "documentation",
        "citation requirements",
        "supplementary information",
    ],
    "Subpopulations": ["counts"],
    "Contacts": ["statement of consent personal data"],
}


def normalize_value(val):
    if pd.isna(val):
        return ""
    return str(val)


def make_target_df(target_sheet: str, target_columns: List[str], source_df: pd.DataFrame | None) -> Tuple[pd.DataFrame, List[str]]:
    out = pd.DataFrame(columns=target_columns)
    unmapped: List[str] = []

    if source_df is None or source_df.empty:
        return out, list(target_columns)

    mapping = COLUMN_MAP.get(target_sheet, {})
    if not mapping:
        return out, list(target_columns)

    # Build rows from source
    rows = []
    for _, src_row in source_df.iterrows():
        row = {col: "" for col in target_columns}
        for src_col, target_col in mapping.items():
            if src_col in source_df.columns and target_col in row:
                row[target_col] = normalize_value(src_row[src_col])
        rows.append(row)

    out = pd.DataFrame(rows, columns=target_columns)
    mapped_targets = set(mapping.values())
    unmapped = [c for c in target_columns if c not in mapped_targets]
    return out, unmapped


def autosize_and_style(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    for idx, col_cells in enumerate(ws.columns, start=1):
        width = max(len(str(c.value or "")) for c in col_cells[:50]) + 2
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(width, 40))


def main():
    parser = argparse.ArgumentParser(description="Convert current extractor workbook to UMCG cohort workbook shape.")
    parser.add_argument("--source", required=True, help="Current extractor output workbook (.xlsx)")
    parser.add_argument("--model", required=True, help="Model workbook such as schemas/molgenis_UMCGCohortsStaging.xlsx")
    parser.add_argument("--output", required=True, help="Converted cohort workbook (.xlsx)")
    parser.add_argument("--report", default=None, help="Optional JSON report path")
    args = parser.parse_args()

    source_path = Path(args.source)
    model_path = Path(args.model)
    output_path = Path(args.output)
    report_path = Path(args.report) if args.report else output_path.with_suffix(".report.json")

    expected = load_model_structure(model_path)
    xls = pd.ExcelFile(source_path)
    source_sheets = {name: pd.read_excel(source_path, sheet_name=name, dtype="object") for name in xls.sheet_names}

    wb = Workbook()
    wb.remove(wb.active)

    report = {
        "source_workbook": str(source_path),
        "model_workbook": str(model_path),
        "output_workbook": str(output_path),
        "mapped_sheets": {},
        "missing_source_sheets": [],
        "known_not_extracted_sheets": KNOWN_NOT_EXTRACTED,
        "partial_columns": KNOWN_PARTIAL_COLUMNS,
        "source_only_sheets": [],
    }

    used_source_sheets = set()

    for target_sheet, target_columns in expected.items():
        source_sheet_name = None
        for src, tgt in SHEET_MAP.items():
            if tgt == target_sheet and src in source_sheets:
                source_sheet_name = src
                used_source_sheets.add(src)
                break

        source_df = source_sheets.get(source_sheet_name) if source_sheet_name else None
        target_df, unmapped = make_target_df(target_sheet, target_columns, source_df)

        ws = wb.create_sheet(title=target_sheet)
        for c_idx, col in enumerate(target_columns, start=1):
            ws.cell(row=1, column=c_idx, value=col)
        for r_idx, (_, row) in enumerate(target_df.iterrows(), start=2):
            for c_idx, col in enumerate(target_columns, start=1):
                ws.cell(row=r_idx, column=c_idx, value=normalize_value(row[col]))
        autosize_and_style(ws)

        report["mapped_sheets"][target_sheet] = {
            "source_sheet": source_sheet_name,
            "row_count": int(len(target_df)),
            "target_column_count": int(len(target_columns)),
            "unmapped_target_columns": unmapped,
        }

        if source_sheet_name is None:
            report["missing_source_sheets"].append(target_sheet)

    report["source_only_sheets"] = [s for s in source_sheets.keys() if s not in used_source_sheets]

    wb.save(output_path)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Saved workbook: {output_path.resolve()}")
    print(f"Saved report:   {report_path.resolve()}")


if __name__ == "__main__":
    main()
