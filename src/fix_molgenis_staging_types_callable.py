from __future__ import annotations

import argparse
import ast
import csv
import difflib
import json
import re
from datetime import date, datetime
from io import StringIO
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook


def _csv_join(items: list[str]) -> str:
    """Comma-join array items with CSV quoting so values containing commas stay intact."""
    if not items:
        return ""
    buf = StringIO()
    csv.writer(buf, lineterminator="").writerow(items)
    return buf.getvalue()

DEFAULT_SCHEMA_WORKBOOK = (
    Path(__file__).resolve().parent.parent / "schemas" / "molgenis_UMCGCohortsStaging.xlsx"
)

# Explicitly cover every column type present in the UMCG staging schema.
ARRAY_TYPES = {"ontology_array", "string_array", "ref_array"}
PASSTHROUGH_TYPES = {"ref", "refback", "ontology", "text"}
INHERITED_TABLE_SCHEMAS: Dict[str, str] = {
    "Organisations": "Agents",
    "Collections": "Resources",
}

REF_FIELD_TARGETS: Dict[tuple[str, str], str] = {
    ("Agents", "resource"): "Resources",
    ("Collection events", "resource"): "Resources",
    ("Organisations", "resource"): "Resources",
    ("Contacts", "organisation"): "Organisations",
    ("Contacts", "resource"): "Resources",
    ("Datasets", "resource"): "Resources",
    ("Documentation", "resource"): "Resources",
    ("External identifiers", "resource"): "Resources",
    ("Internal identifiers", "resource"): "Resources",
    ("Publications", "resource"): "Resources",
    ("Resources", "contact point"): "Contacts",
    ("Resources", "publisher"): "Organisations",
    ("Subpopulation counts", "resource"): "Resources",
    ("Subpopulation counts", "subpopulation"): "Subpopulations",
    ("Subpopulations", "resource"): "Resources",
}

REF_ARRAY_FIELD_TARGETS: Dict[tuple[str, str], str] = {
    ("Collection events", "subpopulations"): "Subpopulations",
    ("Resources", "creator"): "Organisations",
}

# Field-level ontology constraints (first rollout).
ONTOLOGY_ALLOWED_VALUES: Dict[tuple[str, str], set[str]] = {
    ("Agents", "type"): {"Individual", "Organisation"},
    ("Organisations", "type"): {"Individual", "Organisation"},
    ("Resources", "access rights"): {"Open access", "Restricted access", "Non public"},
    ("Resources", "design"): {"Longitudinal", "Cross-sectional"},
    ("Resources", "informed consent type"): {
        "Study specific consent",
        "Broad consent",
        "Passive/tacit consent",
        "No consent",
    },
    ("Resources", "release type"): {
        "Continuous",
        "Closed dataset",
        "Annually",
        "Periodically",
        "Other release type",
    },
    ("Resources", "cohort type"): {
        "Clinical cohort",
        "Population cohort",
        "Case-control",
        "Case only",
        "Birth cohort",
        "Other type",
    },
    ("Resources", "data collection type"): {"Retrospective", "Prospective"},
    ("Collection events", "access rights"): {"Open access", "Restricted access", "Non public"},
    ("Subpopulations", "access rights"): {"Open access", "Restricted access", "Non public"},
    ("Contacts", "title"): {
        "prof. dr.",
        "dr.",
        "ir.",
        "dr. ir.",
        "prof. dr. ir.",
        "prof.",
        "drs.",
    },
    ("Datasets", "unit of observation"): {
        "Drug subscription",
        "person",
        "abortion",
        "Patient",
        "sample",
        "discharge record",
        "medicine dispensation",
        "birth",
        "event",
        "observation",
        "prescription",
        "surgery",
        "treatment",
    },
    ("Documentation", "type"): {
        "Harmonisation protocol",
        "Informed consent",
        "Standard operating protocol",
        "Conflicts of interest of investigators",
        "Composition of steering group and observers",
        "Signed code of conduct",
        "Signed code of conduct checklist",
        "Signed checklist for study protocols",
        "Data characterisation results",
        "Procedure of data extraction",
        "Procedure of results generation",
        "Results tables",
        "Study report",
        "Study, other information",
        "Data source ETL specifications",
        "Governance details",
        "Protocol document",
        "Data characterisation details",
    },
    ("External identifiers", "external identifier type"): {
        "EUPAS number",
        "METc number",
        "Clinical Trials.gov",
        "NCT number",
        "EUDRACT number",
    },
    ("Internal identifiers", "internal identifier type"): {
        "UMCG register Utopia",
        "UMCG PaNaMaID",
    },
    ("Subpopulation counts", "age group"): {
        "Prenatal",
        "All ages",
        "Infant (0-23 months)",
        "Newborn (0-1 months)",
        "Infants and toddlers (2-23 months)",
        "Child (2-12 years)",
        "Adolescent (13-17 years)",
        "Adult (18+ years)",
        "Young adult (18-24 years)",
        "Adult (25-44 years)",
        "Middle-aged (45-64 years)",
        "Aged (65+ years)",
        "Aged (65-79 years)",
        "Aged (80+ years)",
    },
}

ONTOLOGY_ARRAY_ALLOWED_VALUES: Dict[tuple[str, str], set[str]] = {
    ("Resources", "type"): {
        "Biobank",
        "Preclinical study",
        "Clinical trial",
        "Cohort study",
        "Health records",
        "Registry",
        "Reference data source",
        "Common data model",
        "Catalogue",
        "Network",
        "Other type",
    },
    ("Resources", "inclusion criteria"): {
        "Age group inclusion criterion",
        "Age of majority inclusion criterion",
        "BMI range inclusion criterion",
        "Clinically relevant exposure inclusion criterion",
        "Clinically relevant lifestyle inclusion criterion",
        "Country of residence inclusion criteria",
        "Defined population inclusion criterion",
        "Ethnicity inclusion criterion",
        "Family status inclusion criterion",
        "Gravidity inclusion criterion",
        "Health status inclusion criterion",
        "Hospital patient inclusion criterion",
        "Sex inclusion criterion",
        "Use of medication inclusion criterion",
    },
    ("Resources", "exclusion criteria"): {
        "Age group inclusion criterion",
        "Age of majority inclusion criterion",
        "BMI range inclusion criterion",
        "Clinically relevant exposure inclusion criterion",
        "Clinically relevant lifestyle inclusion criterion",
        "Country of residence inclusion criteria",
        "Defined population inclusion criterion",
        "Ethnicity inclusion criterion",
        "Family status inclusion criterion",
        "Gravidity inclusion criterion",
        "Health status inclusion criterion",
        "Hospital patient inclusion criterion",
        "Sex inclusion criterion",
        "Use of medication inclusion criterion",
    },
    ("Resources", "data use conditions"): {
        "research specific restrictions",
        "no general methods research",
        "genetic studies only",
        "not for profit, non commercial use only",
        "publication required",
        "collaboration required",
        "ethics approval required",
        "geographical restriction",
        "publication moratorium",
        "time limit on use",
        "user specific restriction",
        "project specific restriction",
        "institution specific restriction",
        "return to database or resource",
        "clinical care use",
    },
    ("Resources", "data access conditions"): {
        "no restriction",
        "general research use",
        "health or medical or biomedical research",
    },
    ("Datasets", "dataset type"): {
        "Collected dataset",
        "Harmonised dataset",
        "Harmonisation schema",
    },
    ("Contacts", "role"): {
        "Principal Investigator",
        "Primary contact",
        "Project manager",
        "Data manager",
        "Alternative contact",
        "Project leader",
        "Participant",
        "Public lead",
        "EFPIA lead",
        "Task leader",
    },
    ("Collection events", "theme"): {
        "Health",
        "Agriculture",
        "Environment",
        "Energy",
        "Government and public sector",
    },
    ("Resources", "theme"): {
        "Health",
        "Agriculture",
        "Environment",
        "Energy",
        "Government and public sector",
    },
    ("Subpopulations", "theme"): {
        "Health",
        "Agriculture",
        "Environment",
        "Energy",
        "Government and public sector",
    },
    ("Collection events", "data categories"): {
        "Biological samples",
        "Survey data",
        "Imaging data",
        "Medical records",
        "National registries",
        "Genealogical records",
        "Physiological/Biochemical measurements",
        "Omics",
        "Genomics",
        "Epigenomics",
        "Transcriptomics",
        "Proteomics",
        "Metabolomics",
        "Metagenomics / Microbiome",
        "Other",
    },
    ("Collection events", "applicable legislation"): {"Data Governance Act"},
    ("Resources", "applicable legislation"): {"Data Governance Act"},
    ("Subpopulations", "applicable legislation"): {"Data Governance Act"},
    ("Collection events", "age groups"): {
        "Prenatal",
        "All ages",
        "Infant (0-23 months)",
        "Newborn (0-1 months)",
        "Infants and toddlers (2-23 months)",
        "Child (2-12 years)",
        "Adolescent (13-17 years)",
        "Adult (18+ years)",
        "Young adult (18-24 years)",
        "Adult (25-44 years)",
        "Middle-aged (45-64 years)",
        "Aged (65+ years)",
        "Aged (65-79 years)",
        "Aged (80+ years)",
    },
    ("Resources", "population age groups"): {
        "Prenatal",
        "All ages",
        "Infant (0-23 months)",
        "Newborn (0-1 months)",
        "Infants and toddlers (2-23 months)",
        "Child (2-12 years)",
        "Adolescent (13-17 years)",
        "Adult (18+ years)",
        "Young adult (18-24 years)",
        "Adult (25-44 years)",
        "Middle-aged (45-64 years)",
        "Aged (65+ years)",
        "Aged (65-79 years)",
        "Aged (80+ years)",
    },
    ("Subpopulations", "age groups"): {
        "Prenatal",
        "All ages",
        "Infant (0-23 months)",
        "Newborn (0-1 months)",
        "Infants and toddlers (2-23 months)",
        "Child (2-12 years)",
        "Adolescent (13-17 years)",
        "Adult (18+ years)",
        "Young adult (18-24 years)",
        "Adult (25-44 years)",
        "Middle-aged (45-64 years)",
        "Aged (65+ years)",
        "Aged (65-79 years)",
        "Aged (80+ years)",
    },
    ("Subpopulations", "inclusion criteria"): {
        "Age group inclusion criterion",
        "Age of majority inclusion criterion",
        "BMI range inclusion criterion",
        "Clinically relevant exposure inclusion criterion",
        "Clinically relevant lifestyle inclusion criterion",
        "Country of residence inclusion criteria",
        "Defined population inclusion criterion",
        "Ethnicity inclusion criterion",
        "Family status inclusion criterion",
        "Gravidity inclusion criterion",
        "Health status inclusion criterion",
        "Hospital patient inclusion criterion",
        "Sex inclusion criterion",
        "Use of medication inclusion criterion",
    },
    ("Subpopulations", "exclusion criteria"): {
        "Age group inclusion criterion",
        "Age of majority inclusion criterion",
        "BMI range inclusion criterion",
        "Clinically relevant exposure inclusion criterion",
        "Clinically relevant lifestyle inclusion criterion",
        "Country of residence inclusion criteria",
        "Defined population inclusion criterion",
        "Ethnicity inclusion criterion",
        "Family status inclusion criterion",
        "Gravidity inclusion criterion",
        "Health status inclusion criterion",
        "Hospital patient inclusion criterion",
        "Sex inclusion criterion",
        "Use of medication inclusion criterion",
    },
    ("Agents", "role"): {
        "Data originator",
        "Data holder",
        "Data provider",
        "Researcher",
        "Surveillance",
        "Data access provider",
        "Other",
    },
    ("Organisations", "role"): {
        "Data originator",
        "Data holder",
        "Data provider",
        "Researcher",
        "Surveillance",
        "Data access provider",
        "Other",
    },
}

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_DOMAIN_RE = re.compile(
    r"^(?:(?:https?://)|(?:www\.))"
    r"[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    r"(?::\d+)?"
    r"(?:[/?#][^\s]*)?$"
)
_BARE_DOMAIN_RE = re.compile(
    r"^[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    r"(?::\d+)?"
    r"(?:[/?#][^\s]*)?$"
)


# ---------- Schema ----------
def read_schema(schema_path: str, profile: str = "UMCGCohortsStaging") -> Dict[str, Dict[str, Optional[str]]]:
    wb = load_workbook(schema_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(x) if x is not None else "" for x in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}

    schema: Dict[str, Dict[str, Optional[str]]] = {}
    for row in rows:
        table = row[idx.get("tableName", -1)] if "tableName" in idx else None
        col = row[idx.get("columnName", -1)] if "columnName" in idx else None
        col_type = row[idx.get("columnType", -1)] if "columnType" in idx else None
        profiles = row[idx.get("profiles", -1)] if "profiles" in idx else None
        if not table or not col:
            continue
        if profiles and profile not in str(profiles):
            continue
        schema.setdefault(str(table), {})[str(col)] = None if col_type is None else str(col_type)

    for child_table, parent_table in INHERITED_TABLE_SCHEMAS.items():
        if child_table not in schema or parent_table not in schema:
            continue
        merged = dict(schema[parent_table])
        merged.update(schema[child_table])
        schema[child_table] = merged
    return schema


# ---------- Generic helpers ----------
def _clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_ref_token(value: str) -> str:
    s = _clean_string(value).lower()
    if not s:
        return ""
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def parse_list_like(value: Any) -> list[str] | None:
    """
    Convert JSON-like / Python-like list strings to a plain list of strings.
    Returns None when the value is not list-like and should not be touched.
    """
    if value is None:
        return None
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]
    if not isinstance(value, str):
        return None

    s = value.strip()
    if not s:
        return []

    if s.startswith("[") and s.endswith("]"):
        try:
            data = json.loads(s)
            if isinstance(data, list):
                return [str(x).strip() for x in data if str(x).strip()]
        except json.JSONDecodeError:
            pass
        try:
            data = ast.literal_eval(s)
            if isinstance(data, list):
                return [str(x).strip() for x in data if str(x).strip()]
        except (ValueError, SyntaxError):
            pass

    if "," in s:
        return [part.strip() for part in s.split(",") if part.strip()]

    return [s]


def _parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    s = _clean_string(value)
    if not s:
        return None

    s = s.replace("T", " ").replace("Z", "")
    s = re.sub(r"\s+", " ", s)

    known_formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
    ]
    for fmt in known_formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(s)
    except ValueError:
        pass

    m = re.match(r"^(\d{4}-\d{2}-\d{2})(?:[ T](\d{2}:\d{2}(?::\d{2})?))?", s)
    if m:
        date_part = m.group(1)
        time_part = m.group(2) or "00:00:00"
        if len(time_part) == 5:
            time_part += ":00"
        try:
            return datetime.strptime(f"{date_part} {time_part}", "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None

    return None


# ---------- Coercers per schema type ----------
def _parse_array_items(value: Any) -> list[Any] | None:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if not isinstance(value, str):
        return None

    s = value.strip()
    if not s:
        return []

    if s.startswith("[") and s.endswith("]"):
        try:
            data = json.loads(s)
            if isinstance(data, list):
                return data
        except json.JSONDecodeError:
            pass
        try:
            data = ast.literal_eval(s)
            if isinstance(data, list):
                return data
        except (ValueError, SyntaxError):
            pass

    # CSV-style parsing to preserve commas inside quoted values.
    try:
        import csv
        from io import StringIO

        reader = csv.reader(StringIO(s), skipinitialspace=True)
        row = next(reader, None)
        if row is not None and len(row) > 1:
            return [part for part in row]
    except (csv.Error, OSError):
        pass

    if "," in s:
        return [part.strip() for part in s.split(",") if part.strip()]

    return [s]


def coerce_array(table: str, column: str, value: Any) -> Any:
    raw_items = _parse_array_items(value)
    if raw_items is None:
        return value

    items: list[str] = []
    seen: set[str] = set()
    for raw in raw_items:
        s = _extract_ontology_scalar(raw)
        if not s or s in seen:
            continue
        seen.add(s)
        items.append(s)

    allowed = ONTOLOGY_ARRAY_ALLOWED_VALUES.get((table, column))
    if allowed is not None:
        items = [x for x in items if x in allowed]

    return _csv_join(items)


def coerce_bool(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"

    s = _clean_string(value).lower()
    if not s:
        return ""
    if s in {"true", "1", "yes", "y"}:
        return "true"
    if s in {"false", "0", "no", "n"}:
        return "false"
    return ""


def coerce_date(value: Any) -> Any:
    dt = _parse_datetime(value)
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d")


def coerce_datetime(value: Any) -> Any:
    dt = _parse_datetime(value)
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def coerce_email(value: Any) -> Any:
    s = _clean_string(value)
    if not s:
        return ""
    return s if "@" in s and _EMAIL_RE.match(s) else ""


def coerce_file(value: Any) -> Any:
    # User rule: file fields must always stay empty.
    return ""


def coerce_int(value: Any, *, non_negative: bool = False) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        if non_negative and value < 0:
            return ""
        return value
    if isinstance(value, float):
        if value.is_integer():
            ivalue = int(value)
            if non_negative and ivalue < 0:
                return ""
            return ivalue
        return ""

    s = _clean_string(value)
    if not s:
        return ""
    if re.fullmatch(r"[+-]?\d+", s):
        ivalue = int(s)
        if non_negative and ivalue < 0:
            return ""
        return ivalue
    if re.fullmatch(r"[+-]?\d+\.0+", s):
        ivalue = int(float(s))
        if non_negative and ivalue < 0:
            return ""
        return ivalue
    return ""


def coerce_hyperlink(table: str, column: str, value: Any) -> Any:
    s = _clean_string(value)
    if not s:
        return ""

    # Publications.doi is typed as hyperlink in the live EMX2 schema.
    # Accept a bare DOI or DOI URL as input, but always normalize to a
    # canonical DOI resolver URL so downstream imports pass hyperlink validation.
    if table == "Publications" and column == "doi":
        s = re.sub(r"^(?:https?://)?(?:dx\.)?doi\.org/", "", s, flags=re.IGNORECASE)
        s = re.sub(r"^doi:\s*", "", s, flags=re.IGNORECASE)
        s = s.rstrip(".,;:)")
        if re.fullmatch(r"10\.\d{4,9}/[-._;()/:A-Za-z0-9]+", s):
            return f"https://doi.org/{s}"
        return ""

    # User rule: hyperlinks should be a real site, starting with http or www.
    if _DOMAIN_RE.match(s):
        if s.startswith("www."):
            return f"https://{s}"
        return s
    if _BARE_DOMAIN_RE.match(s):
        return f"https://{s}"

    return ""


def coerce_heading(value: Any) -> Any:
    # User rule: heading fields must always be empty.
    return ""


def coerce_passthrough_text(value: Any) -> Any:
    # Keep as plain text, only strip surrounding whitespace.
    return _clean_string(value)


def coerce_passthrough(value: Any) -> Any:
    # For ref / refback / ontology: explicitly handled but currently left unchanged,
    # except for trimming outer whitespace for strings.
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _extract_ref_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        for key in ("id", "pid", "name", "label", "value", "email"):
            s = _clean_string(value.get(key))
            if s:
                return s
        return ""
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return ""
        if s.startswith("{") and s.endswith("}"):
            try:
                parsed = json.loads(s)
            except json.JSONDecodeError:
                return s
            return _extract_ref_scalar(parsed)
        return s
    return _clean_string(value)


def _build_contact_display_name(values: Dict[str, Any]) -> str:
    first = _clean_string(values.get("first name"))
    prefix = _clean_string(values.get("prefix"))
    last = _clean_string(values.get("last name"))
    parts = [p for p in (first, prefix, last) if p]
    return " ".join(parts).strip()


def _sheet_header_index(ws: Any) -> Dict[str, int]:
    headers = [cell.value for cell in ws[1]]
    idx: Dict[str, int] = {}
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue
        idx[str(header)] = col_idx
    return idx


def _sheet_column_values(ws: Any, header_index: Dict[str, int], column_name: str) -> list[str]:
    col_idx = header_index.get(column_name)
    if not col_idx:
        return []
    out: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        s = _clean_string(ws.cell(row=row_idx, column=col_idx).value)
        if s:
            out.append(s)
    return out


def _normalize_csv_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean_string(value).lower())


def _load_external_organisations_index(csv_path: str | Path | None) -> Dict[str, Any]:
    if not csv_path:
        return {}
    path = Path(csv_path)
    if not path.exists():
        return {}

    exact: Dict[str, str] = {}
    exact_multi: set[str] = set()
    norm: Dict[str, str] = {}
    norm_multi: set[str] = set()

    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return {}
        header_map = {
            _normalize_csv_header(field_name): field_name
            for field_name in reader.fieldnames
            if _clean_string(field_name)
        }
        name_field = header_map.get("name")
        if not name_field:
            return {}

        token_fields = [
            header_map.get("name"),
            header_map.get("label"),
            header_map.get("acronym"),
            header_map.get("code"),
            header_map.get("ontologytermuri"),
        ]

        for row in reader:
            canonical = _clean_string(row.get(name_field))
            if not canonical:
                continue

            tokens: set[str] = {canonical}
            for field_name in token_fields:
                if not field_name:
                    continue
                token = _clean_string(row.get(field_name))
                if token:
                    tokens.add(token)

            for token in tokens:
                if token in exact and exact[token] != canonical:
                    exact_multi.add(token)
                else:
                    exact[token] = canonical

                normalized = _normalize_ref_token(token)
                if not normalized:
                    continue
                if normalized in norm and norm[normalized] != canonical:
                    norm_multi.add(normalized)
                else:
                    norm[normalized] = canonical

    for token in exact_multi:
        exact.pop(token, None)
    for token in norm_multi:
        norm.pop(token, None)

    return {
        "exact": exact,
        "norm": norm,
    }


def _match_external_organisation(external_index: Dict[str, Any], raw_value: Any) -> str:
    raw = _clean_string(raw_value)
    if not raw or not external_index:
        return ""

    exact = external_index.get("exact", {})
    if raw in exact:
        return exact[raw]

    normalized = _normalize_ref_token(raw)
    norm = external_index.get("norm", {})
    if normalized and normalized in norm:
        return norm[normalized]

    if not normalized:
        return ""

    score_pairs = [
        (token, difflib.SequenceMatcher(None, normalized, token).ratio())
        for token in norm.keys()
    ]
    score_pairs.sort(key=lambda item: item[1], reverse=True)
    best_token, best_score = score_pairs[0] if score_pairs else ("", 0.0)
    second_score = score_pairs[1][1] if len(score_pairs) > 1 else 0.0
    if best_score >= 0.96 and (best_score - second_score) >= 0.02:
        return norm.get(best_token, "")
    return ""


def _normalize_external_organisation_refs(wb: Any, external_index: Dict[str, Any]) -> None:
    if not external_index:
        return

    for sheet_name in ("Agents", "Organisations"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        hdr = _sheet_header_index(ws)
        organisation_col = hdr.get("organisation")
        other_col = hdr.get("other organisation")
        type_col = hdr.get("type")
        if not organisation_col:
            continue

        for row_idx in range(2, ws.max_row + 1):
            typ = _clean_string(ws.cell(row=row_idx, column=type_col).value) if type_col else ""
            if sheet_name == "Agents" and typ != "Organisation":
                continue
            if sheet_name == "Organisations" and typ and typ != "Organisation":
                continue

            raw_organisation = _clean_string(ws.cell(row=row_idx, column=organisation_col).value)
            raw_other = _clean_string(ws.cell(row=row_idx, column=other_col).value) if other_col else ""
            mapped = ""
            reference_candidates = [raw_organisation]
            if raw_other:
                reference_candidates.append(raw_other)
            for candidate in reference_candidates:
                mapped = _match_external_organisation(external_index, candidate)
                if mapped:
                    break

            if mapped:
                ws.cell(row=row_idx, column=organisation_col).value = mapped
                if other_col:
                    ws.cell(row=row_idx, column=other_col).value = ""
                continue

            if raw_organisation:
                ws.cell(row=row_idx, column=organisation_col).value = ""
                if other_col and not raw_other:
                    ws.cell(row=row_idx, column=other_col).value = raw_organisation


def _build_organisation_ref_maps(wb: Any) -> tuple[Dict[str, str], Dict[str, str]]:
    exact: Dict[str, str] = {}
    norm: Dict[str, str] = {}
    exact_multi: set[str] = set()
    norm_multi: set[str] = set()

    def _add_mapping(token: str, canonical: str) -> None:
        if token in exact and exact[token] != canonical:
            exact_multi.add(token)
        else:
            exact[token] = canonical

        normalized = _normalize_ref_token(token)
        if not normalized:
            return
        if normalized in norm and norm[normalized] != canonical:
            norm_multi.add(normalized)
        else:
            norm[normalized] = canonical

    def _consume(ws: Any, *, require_type_filter: bool) -> None:
        hdr = _sheet_header_index(ws)
        type_col = hdr.get("type")
        id_col = hdr.get("id")
        name_col = hdr.get("name")
        other_col = hdr.get("other organisation")
        for row_idx in range(2, ws.max_row + 1):
            if require_type_filter and type_col:
                typ = _clean_string(ws.cell(row=row_idx, column=type_col).value)
                if typ != "Organisation":
                    continue

            canonical = ""
            for col_idx in (id_col, name_col, other_col):
                if not col_idx:
                    continue
                value = _clean_string(ws.cell(row=row_idx, column=col_idx).value)
                if value:
                    canonical = value
                    break
            if not canonical:
                continue

            for col_idx in (id_col, name_col, other_col):
                if not col_idx:
                    continue
                token = _clean_string(ws.cell(row=row_idx, column=col_idx).value)
                if token:
                    _add_mapping(token, canonical)

    if "Organisations" in wb.sheetnames:
        _consume(wb["Organisations"], require_type_filter=False)
    if "Agents" in wb.sheetnames:
        _consume(wb["Agents"], require_type_filter=True)

    for token in exact_multi:
        exact.pop(token, None)
    for token in norm_multi:
        norm.pop(token, None)
    return exact, norm


def _build_ref_index(wb: Any) -> Dict[str, Dict[str, str]]:
    """
    Build match indices per target table:
    - exact token -> canonical token
    - normalized token -> canonical token (only when unambiguous)
    """
    target_tables = set(REF_FIELD_TARGETS.values())
    candidates: Dict[str, set[str]] = {table: set() for table in target_tables}

    def _collect_sheet_tokens(target_table: str, ws: Any) -> None:
        hdr = _sheet_header_index(ws)
        for col in ("id", "pid", "name", "acronym", "label", "code", "identifier", "email", "organisation", "other organisation"):
            for v in _sheet_column_values(ws, hdr, col):
                candidates[target_table].add(v)

        if target_table == "Contacts":
            for row_idx in range(2, ws.max_row + 1):
                row_values = {
                    key: ws.cell(row=row_idx, column=col_idx).value
                    for key, col_idx in hdr.items()
                }
                display = _build_contact_display_name(row_values)
                if display:
                    candidates[target_table].add(display)

    for table in target_tables:
        if table not in wb.sheetnames:
            continue
        _collect_sheet_tokens(table, wb[table])

    # Include child-table identifiers for parent-table refs (e.g. Collections extends Resources).
    for child_table, parent_table in INHERITED_TABLE_SCHEMAS.items():
        if parent_table not in target_tables:
            continue
        if child_table not in wb.sheetnames:
            continue
        _collect_sheet_tokens(parent_table, wb[child_table])

    # Organisations often extends Agents and may not contain own identifier columns.
    if "Organisations" in target_tables and "Agents" in wb.sheetnames:
        ws_agents = wb["Agents"]
        hdr_agents = _sheet_header_index(ws_agents)
        type_col = hdr_agents.get("type")
        id_col = hdr_agents.get("id")
        name_col = hdr_agents.get("name")
        organisation_col = hdr_agents.get("organisation")
        if type_col:
            for row_idx in range(2, ws_agents.max_row + 1):
                typ = _clean_string(ws_agents.cell(row=row_idx, column=type_col).value)
                if typ != "Organisation":
                    continue
                if id_col:
                    id_val = _clean_string(ws_agents.cell(row=row_idx, column=id_col).value)
                    if id_val:
                        candidates["Organisations"].add(id_val)
                if name_col:
                    name_val = _clean_string(ws_agents.cell(row=row_idx, column=name_col).value)
                    if name_val:
                        candidates["Organisations"].add(name_val)
                if organisation_col:
                    organisation_val = _clean_string(ws_agents.cell(row=row_idx, column=organisation_col).value)
                    if organisation_val:
                        candidates["Organisations"].add(organisation_val)

    ref_index: Dict[str, Dict[str, str]] = {}
    organisation_exact, organisation_norm = _build_organisation_ref_maps(wb)
    for table, vals in candidates.items():
        if table == "Organisations" and (organisation_exact or organisation_norm):
            canonical_values = set(organisation_exact.values()) | set(organisation_norm.values())
            ref_index[table] = {"__size__": str(len(canonical_values))}
            ref_index[table].update({f"e:{k}": v for k, v in organisation_exact.items()})
            ref_index[table].update({f"n:{k}": v for k, v in organisation_norm.items()})
            continue

        by_exact: Dict[str, str] = {}
        by_norm: Dict[str, str] = {}
        norm_multi: set[str] = set()
        for v in vals:
            by_exact[v] = v
            n = _normalize_ref_token(v)
            if not n:
                continue
            if n in by_norm and by_norm[n] != v:
                norm_multi.add(n)
            else:
                by_norm[n] = v
        for n in norm_multi:
            by_norm.pop(n, None)
        ref_index[table] = {"__size__": str(len(vals))}
        ref_index[table].update({f"e:{k}": v for k, v in by_exact.items()})
        ref_index[table].update({f"n:{k}": v for k, v in by_norm.items()})

    return ref_index


def coerce_ref(table: str, column: str, value: Any, ref_index: Dict[str, Dict[str, str]] | None) -> Any:
    s = _extract_ref_scalar(value)
    if not s:
        return ""

    target_table = REF_FIELD_TARGETS.get((table, column))
    if not target_table or not ref_index:
        return s
    target_idx = ref_index.get(target_table)
    if not target_idx:
        return ""

    mapped = _resolve_ref_token(target_idx, s)
    if mapped:
        return mapped
    return ""


def _resolve_ref_token(target_idx: Dict[str, str], token: str) -> str:
    exact = target_idx.get(f"e:{token}")
    if exact:
        return exact
    n = _normalize_ref_token(token)
    if n:
        mapped = target_idx.get(f"n:{n}")
        if mapped:
            return mapped
    return ""

def coerce_ref_array(table: str, column: str, value: Any, ref_index: Dict[str, Dict[str, str]] | None) -> Any:
    raw_items = _parse_array_items(value)
    if raw_items is None:
        return value

    items: list[str] = []
    seen: set[str] = set()
    for raw in raw_items:
        s = _extract_ref_scalar(raw)
        if not s:
            continue
        if s in seen:
            continue
        seen.add(s)
        items.append(s)

    target_table = REF_ARRAY_FIELD_TARGETS.get((table, column))
    if not target_table or not ref_index:
        return _csv_join(items)

    target_idx = ref_index.get(target_table)
    if not target_idx:
        return ""

    resolved: list[str] = []
    resolved_seen: set[str] = set()
    for item in items:
        mapped = _resolve_ref_token(target_idx, item)
        if not mapped or mapped in resolved_seen:
            continue
        resolved_seen.add(mapped)
        resolved.append(mapped)
    return _csv_join(resolved)


def _extract_ontology_scalar(value: Any) -> str:
    """
    Accept common ontology payload variants and return a scalar value:
    - "Organisation"
    - {"name": "Organisation"}
    - "{\"name\":\"Organisation\"}"
    """
    if value is None:
        return ""

    if isinstance(value, dict):
        for key in ("name", "id", "label", "value"):
            v = value.get(key)
            s = _clean_string(v)
            if s:
                return s
        return ""

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return ""

        # Try to parse serialized structured payloads first.
        if (s.startswith("{") and s.endswith("}")) or (s.startswith("[") and s.endswith("]")):
            try:
                parsed = json.loads(s)
                if isinstance(parsed, list):
                    for item in parsed:
                        scalar = _extract_ontology_scalar(item)
                        if scalar:
                            return scalar
                    return ""
                return _extract_ontology_scalar(parsed)
            except json.JSONDecodeError:
                pass

        # Handle partially broken wrappers, e.g. 'Data access provider"]' or '["Data provider'.
        s = re.sub(r'^[\[\]"\\\']+', "", s)
        s = re.sub(r'[\[\]"\\\']+$', "", s)
        return _clean_string(s)

    return _clean_string(value)


def coerce_ontology(table: str, column: str, value: Any) -> Any:
    s = _extract_ontology_scalar(value)
    if not s:
        return ""

    allowed = ONTOLOGY_ALLOWED_VALUES.get((table, column))
    if allowed is None:
        return s
    if s in allowed:
        return s
    return ""


# ---------- Main normalization ----------
def normalize_value(
    table: str,
    column: str,
    col_type: Optional[str],
    value: Any,
    ref_index: Dict[str, Dict[str, str]] | None = None,
) -> Any:
    if col_type == "ref_array":
        return coerce_ref_array(table, column, value, ref_index)
    if col_type in ARRAY_TYPES:
        return coerce_array(table, column, value)
    if col_type == "ref":
        return coerce_ref(table, column, value, ref_index)
    if col_type == "refback":
        return ""
    if col_type == "heading":
        return coerce_heading(value)
    if col_type == "date":
        return coerce_date(value)
    if col_type == "datetime":
        return coerce_datetime(value)
    if col_type == "email":
        return coerce_email(value)
    if col_type == "file":
        return coerce_file(value)
    if col_type == "int":
        return coerce_int(value, non_negative=False)
    if col_type == "non_negative_int":
        return coerce_int(value, non_negative=True)
    if col_type == "hyperlink":
        return coerce_hyperlink(table, column, value)
    if col_type == "bool":
        return coerce_bool(value)
    if col_type == "text":
        return coerce_passthrough_text(value)
    if col_type == "ontology":
        return coerce_ontology(table, column, value)
    if col_type in PASSTHROUGH_TYPES:
        return coerce_passthrough(value)
    return value


# ---------- Workbook runner ----------
def fix_workbook(
    input_path: str | Path,
    schema_path: str | Path = DEFAULT_SCHEMA_WORKBOOK,
    output_path: str | Path | None = None,
    profile: str = "UMCGCohortsStaging",
    ref_organisations_csv: str | Path | None = None,
) -> Path:
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else input_path.with_name(input_path.stem + "_fixed.xlsx")

    schema = read_schema(str(schema_path), profile=profile)
    wb = load_workbook(input_path)

    # Pass 1: normalize all non-ref columns.
    for sheet_name in wb.sheetnames:
        if sheet_name not in schema:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            column_name = str(header)
            col_type = schema[sheet_name].get(column_name)
            if not col_type:
                continue
            if col_type in {"ref", "refback"}:
                continue
            if col_type == "ref_array" and (sheet_name, column_name) in REF_ARRAY_FIELD_TARGETS:
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = normalize_value(sheet_name, column_name, col_type, cell.value)

    _normalize_external_organisation_refs(
        wb,
        _load_external_organisations_index(ref_organisations_csv),
    )

    # Pass 2: validate/normalize refs against current workbook values.
    ref_index = _build_ref_index(wb)
    for sheet_name in wb.sheetnames:
        if sheet_name not in schema:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            column_name = str(header)
            col_type = schema[sheet_name].get(column_name)
            if col_type == "ref_array" and (sheet_name, column_name) in REF_ARRAY_FIELD_TARGETS:
                pass
            elif col_type not in {"ref", "refback"}:
                continue
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = normalize_value(
                    sheet_name,
                    column_name,
                    col_type,
                    cell.value,
                    ref_index=ref_index,
                )

    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize MOLGENIS staging workbook cell values using schema types.")
    parser.add_argument("input", help="Input workbook to fix")
    parser.add_argument("-s", "--schema", default=str(DEFAULT_SCHEMA_WORKBOOK), help="Schema workbook")
    parser.add_argument("-o", "--output", default=None, help="Output workbook path")
    parser.add_argument("--profile", default="UMCGCohortsStaging", help="Profile name in schema workbook")
    parser.add_argument("--ref-organisations-csv", default=None, help="Optional Organisations.csv ontology file")
    args = parser.parse_args()

    output_path = fix_workbook(
        input_path=args.input,
        schema_path=args.schema,
        output_path=args.output,
        profile=args.profile,
        ref_organisations_csv=args.ref_organisations_csv,
    )
    print(output_path)


if __name__ == "__main__":
    main()
