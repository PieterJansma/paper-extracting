from __future__ import annotations

import argparse
import ast
import csv
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional
import difflib

from openpyxl import load_workbook
from llm_client import OpenAICompatibleClient


def _strip_accents(text: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch)
    )


def _normalize_key(value: str) -> str:
    s = _strip_accents(value).lower().strip()
    s = s.replace("&", " and ")
    s = re.sub(r"\(the\)", "", s)
    s = re.sub(r"[\(\)\[\]\*'’`\".,;/_-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_list_like(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]

    s = str(value).strip()
    if not s:
        return []

    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass

    if "," in s:
        return [part.strip() for part in s.split(",") if part.strip()]
    return [s]


@dataclass
class MatchResult:
    raw: str
    mapped: Optional[str]
    method: str
    suggestions: List[str]


class CountryMapper:
    """
    Deterministic country mapper against a Countries ontology CSV.
    """

    # Common aliases/synonyms to canonical ontology names.
    _ALIASES: Dict[str, str] = {
        "the netherlands": "Netherlands (the)",
        "netherlands": "Netherlands (the)",
        "holland": "Netherlands (the)",
        "uk": "United Kingdom of Great Britain and Northern Ireland (the)",
        "u k": "United Kingdom of Great Britain and Northern Ireland (the)",
        "united kingdom": "United Kingdom of Great Britain and Northern Ireland (the)",
        "great britain": "United Kingdom of Great Britain and Northern Ireland (the)",
        "england": "United Kingdom of Great Britain and Northern Ireland (the)",
        "us": "United States of America (the)",
        "u s": "United States of America (the)",
        "usa": "United States of America (the)",
        "united states": "United States of America (the)",
        "united states of america": "United States of America (the)",
        "czech republic": "Czechia",
        "ivory coast": "Côte d'Ivoire",
        "south korea": "Korea (the Republic of)",
        "north korea": "Korea (the Democratic People's Republic of)",
        "russia": "Russian Federation (the)",
        "laos": "Lao People's Democratic Republic (the)",
        "syria": "Syrian Arab Republic (the)",
        "iran": "Iran (Islamic Republic of)",
        "moldova": "Moldova (the Republic of)",
        "venezuela": "Venezuela (Bolivarian Republic of)",
        "tanzania": "Tanzania, the United Republic of",
        "palestine": "Palestine, State of",
        "vietnam": "Viet Nam",
        "brunei": "Brunei Darussalam",
    }

    def __init__(
        self,
        ontology_csv: Path,
        *,
        fuzzy_cutoff: float = 0.96,
        fuzzy_margin: float = 0.02,
    ) -> None:
        self.ontology_csv = ontology_csv
        self.fuzzy_cutoff = fuzzy_cutoff
        self.fuzzy_margin = fuzzy_margin
        self._canonical_names: set[str] = set()
        self._by_code: Dict[str, str] = {}
        self._by_lower_name: Dict[str, str] = {}
        self._by_norm_name: Dict[str, str] = {}
        self._load()

    def _load(self) -> None:
        with self.ontology_csv.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = str(row.get("name", "")).strip()
                if not name:
                    continue
                code = str(row.get("code", "")).strip().upper()
                self._canonical_names.add(name)
                self._by_lower_name[name.lower()] = name
                norm = _normalize_key(name)
                self._by_norm_name.setdefault(norm, name)
                # Alternate normalized key with leading "the " removed.
                norm_no_the = re.sub(r"^the\s+", "", norm).strip()
                if norm_no_the:
                    self._by_norm_name.setdefault(norm_no_the, name)
                if code:
                    self._by_code[code] = name

    def match_one(self, raw_value: str) -> MatchResult:
        raw = str(raw_value or "").strip()
        if not raw:
            return MatchResult(raw=raw, mapped=None, method="empty", suggestions=[])

        upper = raw.upper()
        if upper in self._by_code:
            return MatchResult(raw=raw, mapped=self._by_code[upper], method="code", suggestions=[])

        low = raw.lower()
        if low in self._by_lower_name:
            return MatchResult(raw=raw, mapped=self._by_lower_name[low], method="exact_name", suggestions=[])

        norm = _normalize_key(raw)
        alias_target = self._ALIASES.get(norm)
        if alias_target and alias_target in self._canonical_names:
            return MatchResult(raw=raw, mapped=alias_target, method="alias", suggestions=[])

        if norm in self._by_norm_name:
            return MatchResult(raw=raw, mapped=self._by_norm_name[norm], method="normalized_name", suggestions=[])

        keys = list(self._by_norm_name.keys())
        score_pairs = [
            (k, difflib.SequenceMatcher(None, norm, k).ratio())
            for k in keys
        ]
        score_pairs.sort(key=lambda x: x[1], reverse=True)
        top_key, top_score = score_pairs[0] if score_pairs else ("", 0.0)
        second_score = score_pairs[1][1] if len(score_pairs) > 1 else 0.0

        if top_score >= self.fuzzy_cutoff and (top_score - second_score) >= self.fuzzy_margin:
            return MatchResult(
                raw=raw,
                mapped=self._by_norm_name[top_key],
                method="fuzzy_auto",
                suggestions=[],
            )

        # Keep suggestions for optional LLM fallback.
        seen: set[str] = set()
        suggestions: List[str] = []
        for k, _ in score_pairs[:5]:
            cand = self._by_norm_name[k]
            if cand in seen:
                continue
            seen.add(cand)
            suggestions.append(cand)
        return MatchResult(raw=raw, mapped=None, method="unmapped", suggestions=suggestions)

    @staticmethod
    def _resolve_with_llm(
        raw: str,
        candidates: List[str],
        client: OpenAICompatibleClient,
    ) -> Optional[str]:
        if not candidates:
            return None
        system = (
            "You map country mentions to one canonical country name. "
            "Return ONLY JSON: {\"choice\": <candidate or null>}."
        )
        user = (
            f"Raw value: {raw}\n"
            f"Candidates: {json.dumps(candidates, ensure_ascii=False)}\n"
            "Choose exactly one candidate if clearly matching, otherwise null."
        )
        try:
            text = client.chat(
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                temperature=0.0,
                max_tokens=256,
                response_format={"type": "json_object"},
                max_retries=2,
            )
            payload = json.loads(text.strip())
            choice = payload.get("choice")
            if isinstance(choice, str) and choice in candidates:
                return choice
        except Exception:
            return None
        return None

    def map_list(
        self,
        values: List[str],
        *,
        llm_client: Optional[OpenAICompatibleClient] = None,
        llm_max_candidates: int = 5,
        llm_max_lookups: int = 50,
    ) -> Dict[str, Any]:
        mapped: List[str] = []
        unmapped: List[Dict[str, Any]] = []
        matches: List[Dict[str, Any]] = []
        seen: set[str] = set()
        mapped_input_count = 0
        llm_lookups = 0

        for raw in values:
            res = self.match_one(raw)
            if res.mapped:
                mapped_input_count += 1
                if res.mapped not in seen:
                    seen.add(res.mapped)
                    mapped.append(res.mapped)
                matches.append(
                    {
                        "raw": res.raw,
                        "mapped": res.mapped,
                        "method": res.method,
                    }
                )
            else:
                llm_choice: Optional[str] = None
                if (
                    llm_client is not None
                    and llm_lookups < llm_max_lookups
                    and res.suggestions
                ):
                    llm_lookups += 1
                    llm_choice = self._resolve_with_llm(
                        raw=res.raw,
                        candidates=res.suggestions[:llm_max_candidates],
                        client=llm_client,
                    )
                if llm_choice:
                    mapped_input_count += 1
                    if llm_choice not in seen:
                        seen.add(llm_choice)
                        mapped.append(llm_choice)
                    matches.append(
                        {
                            "raw": res.raw,
                            "mapped": llm_choice,
                            "method": "llm_candidate",
                            "suggestions": res.suggestions,
                        }
                    )
                    continue
                matches.append(
                    {
                        "raw": res.raw,
                        "mapped": None,
                        "method": res.method,
                        "suggestions": res.suggestions,
                    }
                )
                unmapped.append(
                    {
                        "raw": res.raw,
                        "method": res.method,
                        "suggestions": res.suggestions,
                    }
                )

        return {
            "input_count": len(values),
            "mapped_input_count": mapped_input_count,
            "mapped_unique_count": len(mapped),
            "unmapped_count": len(unmapped),
            "mapped": mapped,
            "matches": matches,
            "unmapped": unmapped,
        }


def _collect_values(args: argparse.Namespace) -> List[str]:
    out: List[str] = []
    for v in args.value or []:
        out.extend(_parse_list_like(v))

    if args.values_file:
        p = Path(args.values_file)
        text = p.read_text(encoding="utf-8")
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            out.extend(_parse_list_like(line))

    if args.values_json:
        parsed = json.loads(args.values_json)
        if not isinstance(parsed, list):
            raise SystemExit("--values-json must be a JSON array")
        out.extend([str(x).strip() for x in parsed if str(x).strip()])

    return out


def _map_csv_column(
    mapper: CountryMapper,
    input_csv: Path,
    output_csv: Path,
    column: str,
    issues_json: Optional[Path],
    llm_client: Optional[OpenAICompatibleClient] = None,
) -> None:
    with input_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = list(reader.fieldnames or [])

    if column not in fieldnames:
        raise SystemExit(f"Column not found in input CSV: {column}")

    issues: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        raw = row.get(column, "")
        values = _parse_list_like(raw)
        res = mapper.map_list(values, llm_client=llm_client)
        row[column] = ",".join(res["mapped"])
        for item in res["unmapped"]:
            issues.append(
                {
                    "row": idx,
                    "column": column,
                    "raw": item["raw"],
                    "suggestions": item["suggestions"],
                }
            )

    with output_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    if issues_json:
        issues_json.write_text(json.dumps({"issues": issues}, ensure_ascii=False, indent=2), encoding="utf-8")


def map_workbook_countries(
    workbook_path: Path,
    ontology_csv: Path,
    *,
    output_path: Optional[Path] = None,
    issues_json: Optional[Path] = None,
    llm_client: Optional[OpenAICompatibleClient] = None,
) -> Dict[str, Any]:
    """
    Map country values for known workbook columns:
    - Resources.countries
    - Subpopulations.countries
    """
    mapper = CountryMapper(ontology_csv)
    wb = load_workbook(workbook_path)

    targets = [
        ("Resources", "countries"),
        ("Subpopulations", "countries"),
    ]

    issues: List[Dict[str, Any]] = []
    mapped_cells = 0
    touched_sheets = 0

    for sheet_name, column_name in targets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_idx = None
        for idx, h in enumerate(headers, start=1):
            if str(h or "").strip() == column_name:
                col_idx = idx
                break
        if col_idx is None:
            continue
        touched_sheets += 1

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            raw = cell.value
            values = _parse_list_like(raw)
            res = mapper.map_list(values, llm_client=llm_client)
            new_value = ",".join(res["mapped"])
            if str(cell.value or "") != new_value:
                mapped_cells += 1
            cell.value = new_value

            for item in res["unmapped"]:
                issues.append(
                    {
                        "sheet": sheet_name,
                        "row": row_idx,
                        "column": column_name,
                        "raw": item["raw"],
                        "suggestions": item["suggestions"],
                    }
                )

    save_path = output_path or workbook_path
    wb.save(save_path)

    if issues_json:
        issues_json.write_text(
            json.dumps({"issues": issues}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return {
        "workbook": str(save_path),
        "touched_sheets": touched_sheets,
        "mapped_cells": mapped_cells,
        "issue_count": len(issues),
    }


def cli() -> None:
    parser = argparse.ArgumentParser(
        description="Map country values to canonical ontology names from Countries.csv"
    )
    parser.add_argument("--ontology-csv", required=True, help="Path to Countries.csv ontology file")
    parser.add_argument("--value", action="append", help="Single value or list-like value to map")
    parser.add_argument("--values-file", help="Text file with one value (or list-like value) per line")
    parser.add_argument("--values-json", help="JSON array of values")
    parser.add_argument("--output-json", help="Optional output JSON file for mapped/unmapped results")

    parser.add_argument("--input-csv", help="Optional input CSV to map in-place by column")
    parser.add_argument("--column", help="Column in --input-csv containing country values")
    parser.add_argument("--output-csv", help="Output CSV path for --input-csv mode")
    parser.add_argument("--issues-json", help="Optional issues report path for --input-csv mode")
    parser.add_argument("--input-xlsx", help="Optional input workbook to map country columns")
    parser.add_argument("--output-xlsx", help="Output workbook path for --input-xlsx mode")
    parser.add_argument("--llm-fallback", action="store_true", help="Enable LLM fallback for unresolved values")
    parser.add_argument("--llm-base-url", default="http://127.0.0.1:8080/v1", help="LLM base URL (OpenAI-compatible)")
    parser.add_argument("--llm-api-key", default="sk-local", help="LLM API key")
    parser.add_argument("--llm-model", default="numind/NuExtract-2.0-8B", help="LLM model")

    args = parser.parse_args()
    mapper = CountryMapper(Path(args.ontology_csv))
    llm_client: Optional[OpenAICompatibleClient] = None
    if args.llm_fallback:
        llm_client = OpenAICompatibleClient(
            base_url=args.llm_base_url,
            api_key=args.llm_api_key,
            model=args.llm_model,
            use_grammar=False,
            use_session=False,
        )

    if args.input_xlsx:
        result = map_workbook_countries(
            workbook_path=Path(args.input_xlsx),
            ontology_csv=Path(args.ontology_csv),
            output_path=Path(args.output_xlsx) if args.output_xlsx else None,
            issues_json=Path(args.issues_json) if args.issues_json else None,
            llm_client=llm_client,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    if args.input_csv:
        if not args.column or not args.output_csv:
            raise SystemExit("--input-csv requires --column and --output-csv")
        _map_csv_column(
            mapper=mapper,
            input_csv=Path(args.input_csv),
            output_csv=Path(args.output_csv),
            column=args.column,
            issues_json=Path(args.issues_json) if args.issues_json else None,
            llm_client=llm_client,
        )
        print(Path(args.output_csv).resolve())
        return

    values = _collect_values(args)
    result = mapper.map_list(values, llm_client=llm_client)
    payload = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output_json:
        Path(args.output_json).write_text(payload, encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    cli()
