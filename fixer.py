from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

INPUT_WORKBOOK_NAME = "alle_data_tabs_leesbaar.xlsx"
SELECTION_WORKBOOK_NAME = "selectie patienten.xlsx"
KEY_FILE_NAME = "sleutel_final(Sheet1).csv"

AZGNR_COLUMN = "AZGNR"
PATIENT_IDENTIFIER_COLUMN = "PatientIdentifier"
STUDY_ID_NEW_COLUMN = "studie_id_new"
STUDY_ID_COLUMN = "Study subject identifier"
SUBJECT_ID_COLUMN = "Subject identifier (technical)"
DOB_COLUMN = "DOB"
START_DATE_COLUMN = "startdt"
PRIMARY_SUBJECT_LOOKUP_SHEET = "data_NL_CM_0_1_1"

SELECTION_OUTPUT_SHEET = "selection_filtered"
SUMMARY_OUTPUT_SHEET = "filter_summary"


def script_dir() -> Path:
    return Path(__file__).resolve().parent


def default_input_path() -> Path:
    return script_dir() / INPUT_WORKBOOK_NAME


def default_selection_path() -> Path:
    return script_dir() / SELECTION_WORKBOOK_NAME


def default_key_path() -> Path:
    return script_dir() / KEY_FILE_NAME


def default_output_path(input_path: Path, min_age: int, max_age: int) -> Path:
    return input_path.with_name(
        f"{input_path.stem}_selectie_leeftijd_{min_age}_{max_age}.xlsx"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Map selectiepatienten via de sleutellijst naar studie_id_new, "
            "filter op leeftijd bij start behandeling, en pas die selectie toe "
            "op alle tabs van een Excel-werkmap."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=default_input_path(),
        help="Bronbestand met alle leesbare data tabs (.xlsx).",
    )
    parser.add_argument(
        "--selection",
        type=Path,
        default=default_selection_path(),
        help="Excelbestand met de patiëntselectie (.xlsx).",
    )
    parser.add_argument(
        "--selection-sheet",
        type=str,
        default=None,
        help="Optionele naam van het tabblad in het selectie-excelbestand.",
    )
    parser.add_argument(
        "--key",
        type=Path,
        default=default_key_path(),
        help="CSV met de mapping PatientIdentifier -> studie_id_new.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Gefilterde output-werkmap (.xlsx). "
            "Standaard: <input>_selectie_leeftijd_<min>_<max>.xlsx."
        ),
    )
    parser.add_argument("--min-age", type=int, default=18, help="Minimumleeftijd, inclusief.")
    parser.add_argument("--max-age", type=int, default=50, help="Maximumleeftijd, inclusief.")
    return parser.parse_args()


def normalize_value(value: object) -> str | None:
    if pd.isna(value):
        return None

    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip() or None


def detect_identifier_width(series: pd.Series) -> int | None:
    lengths = series.map(normalize_value).dropna().map(len)
    if lengths.empty:
        return None
    return int(lengths.mode().iloc[0])


def normalize_identifier(value: object, width: int | None = None) -> str | None:
    normalized = normalize_value(value)
    if normalized is None:
        return None

    if normalized.endswith(".0") and normalized.replace(".0", "", 1).isdigit():
        normalized = normalized[:-2]

    if width is not None and normalized.isdigit():
        normalized = normalized.zfill(width)

    return normalized


def normalize_series(series: pd.Series, width: int | None = None) -> pd.Series:
    return series.map(lambda value: normalize_identifier(value, width=width))


def calculate_age_at_start(dob: pd.Series, start_date: pd.Series) -> pd.Series:
    before_birthday = (start_date.dt.month < dob.dt.month) | (
        (start_date.dt.month == dob.dt.month) & (start_date.dt.day < dob.dt.day)
    )
    return (
        start_date.dt.year
        - dob.dt.year
        - before_birthday.fillna(False).astype(int)
    )


def require_columns(frame: pd.DataFrame, required_columns: list[str], source_name: str) -> None:
    missing_columns = [column for column in required_columns if column not in frame.columns]
    if missing_columns:
        raise SystemExit(
            f"Missing column(s) in {source_name}: {', '.join(missing_columns)}"
        )


def read_selection_frame(path: Path, sheet_name: str | None) -> tuple[pd.DataFrame, str]:
    workbook = pd.ExcelFile(path)
    selected_sheet_name = sheet_name or workbook.sheet_names[0]
    if selected_sheet_name not in workbook.sheet_names:
        raise SystemExit(
            f"Selection sheet '{selected_sheet_name}' not found in {path.name}. "
            f"Available sheets: {', '.join(workbook.sheet_names)}"
        )

    frame = pd.read_excel(path, sheet_name=selected_sheet_name, dtype="object")
    require_columns(
        frame,
        [AZGNR_COLUMN, DOB_COLUMN, START_DATE_COLUMN],
        f"{path.name}:{selected_sheet_name}",
    )
    return frame, selected_sheet_name


def read_key_frame(path: Path) -> pd.DataFrame:
    frame = pd.read_csv(path, dtype="object")
    require_columns(
        frame,
        [PATIENT_IDENTIFIER_COLUMN, STUDY_ID_NEW_COLUMN],
        path.name,
    )
    return frame


def prepare_selection(
    selection_frame: pd.DataFrame,
    key_frame: pd.DataFrame,
    min_age: int,
    max_age: int,
) -> tuple[pd.DataFrame, pd.DataFrame, int]:
    patient_id_width = detect_identifier_width(key_frame[PATIENT_IDENTIFIER_COLUMN])

    key_map = key_frame[[PATIENT_IDENTIFIER_COLUMN, STUDY_ID_NEW_COLUMN]].copy()
    key_map["patient_identifier_normalized"] = normalize_series(
        key_map[PATIENT_IDENTIFIER_COLUMN], width=patient_id_width
    )
    key_map[STUDY_ID_NEW_COLUMN] = key_map[STUDY_ID_NEW_COLUMN].map(normalize_value)
    key_map = key_map.dropna(
        subset=["patient_identifier_normalized", STUDY_ID_NEW_COLUMN]
    ).copy()

    conflicts = (
        key_map.groupby("patient_identifier_normalized")[STUDY_ID_NEW_COLUMN]
        .nunique()
        .loc[lambda series: series > 1]
    )
    if not conflicts.empty:
        raise SystemExit(
            "Conflicting mappings found in key file for PatientIdentifier values: "
            + ", ".join(conflicts.index.tolist()[:10])
        )

    key_map = key_map.drop_duplicates(
        subset=["patient_identifier_normalized"], keep="first"
    )

    selection = selection_frame.copy()
    selection["azgnr_normalized"] = normalize_series(
        selection[AZGNR_COLUMN], width=patient_id_width
    )
    selection = selection.merge(
        key_map[["patient_identifier_normalized", STUDY_ID_NEW_COLUMN]],
        how="left",
        left_on="azgnr_normalized",
        right_on="patient_identifier_normalized",
    )

    selection[DOB_COLUMN] = pd.to_datetime(selection[DOB_COLUMN], errors="coerce")
    selection[START_DATE_COLUMN] = pd.to_datetime(
        selection[START_DATE_COLUMN], errors="coerce"
    )
    selection["age_at_start"] = calculate_age_at_start(
        selection[DOB_COLUMN], selection[START_DATE_COLUMN]
    )
    selection["has_study_id_new"] = selection[STUDY_ID_NEW_COLUMN].notna()
    selection["within_age_range"] = selection["age_at_start"].between(
        min_age, max_age, inclusive="both"
    )
    selection["selected_for_output"] = (
        selection["has_study_id_new"] & selection["within_age_range"]
    )

    filtered_selection = selection.loc[selection["selected_for_output"]].copy()
    return selection, filtered_selection, patient_id_width or 0


def build_column_index(header: list[object] | tuple[object, ...]) -> dict[str, int]:
    return {
        str(column_name): index
        for index, column_name in enumerate(header)
        if column_name is not None
    }


def extract_eligible_subject_ids_from_sheet(
    worksheet,
    eligible_study_ids: set[str],
) -> set[str]:
    row_iterator = worksheet.iter_rows(values_only=True)
    try:
        header = next(row_iterator)
    except StopIteration:
        return set()

    column_index = build_column_index(header)
    study_index = column_index.get(STUDY_ID_COLUMN)
    subject_index = column_index.get(SUBJECT_ID_COLUMN)
    if study_index is None or subject_index is None:
        return set()

    eligible_subject_ids: set[str] = set()
    for row in row_iterator:
        study_id = normalize_identifier(row[study_index])
        if study_id not in eligible_study_ids:
            continue

        subject_id = normalize_identifier(row[subject_index])
        if subject_id is not None:
            eligible_subject_ids.add(subject_id)

    return eligible_subject_ids


def build_eligible_subject_ids(path: Path, eligible_study_ids: set[str]) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if PRIMARY_SUBJECT_LOOKUP_SHEET in workbook.sheetnames:
            subject_ids = extract_eligible_subject_ids_from_sheet(
                workbook[PRIMARY_SUBJECT_LOOKUP_SHEET],
                eligible_study_ids,
            )
            if subject_ids:
                return subject_ids

        eligible_subject_ids: set[str] = set()
        for worksheet in workbook.worksheets:
            eligible_subject_ids.update(
                extract_eligible_subject_ids_from_sheet(worksheet, eligible_study_ids)
            )
        return eligible_subject_ids
    finally:
        workbook.close()


def output_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def append_dataframe_sheet(
    workbook: Workbook,
    sheet_name: str,
    frame: pd.DataFrame,
) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(frame.columns.tolist())
    for row in frame.itertuples(index=False, name=None):
        worksheet.append([output_value(value) for value in row])


def stream_filter_workbook(
    input_path: Path,
    output_path: Path,
    eligible_study_ids: set[str],
    eligible_subject_ids: set[str],
    selection_output: pd.DataFrame,
) -> list[tuple[str, int, int, str]]:
    input_workbook = load_workbook(input_path, read_only=True, data_only=True)
    output_workbook = Workbook(write_only=True)
    sheet_report: list[tuple[str, int, int, str]] = []

    try:
        for input_worksheet in input_workbook.worksheets:
            print(f"Processing sheet: {input_worksheet.title}", flush=True)

            output_worksheet = output_workbook.create_sheet(title=input_worksheet.title)
            row_iterator = input_worksheet.iter_rows(values_only=True)

            try:
                header = next(row_iterator)
            except StopIteration:
                output_worksheet.append([])
                sheet_report.append((input_worksheet.title, 0, 0, "unfiltered"))
                continue

            header_list = list(header)
            output_worksheet.append(header_list)

            column_index = build_column_index(header_list)
            study_index = column_index.get(STUDY_ID_COLUMN)
            subject_index = column_index.get(SUBJECT_ID_COLUMN)
            total_rows = 0
            kept_rows = 0

            if study_index is not None and subject_index is not None:
                filter_key = f"{STUDY_ID_COLUMN} or {SUBJECT_ID_COLUMN}"
            elif study_index is not None:
                filter_key = STUDY_ID_COLUMN
            elif subject_index is not None:
                filter_key = SUBJECT_ID_COLUMN
            else:
                filter_key = "unfiltered"

            for row in row_iterator:
                total_rows += 1

                if study_index is None and subject_index is None:
                    keep_row = True
                else:
                    keep_row = False
                    if study_index is not None:
                        study_id = normalize_identifier(row[study_index])
                        keep_row = study_id in eligible_study_ids
                    if not keep_row and subject_index is not None:
                        subject_id = normalize_identifier(row[subject_index])
                        keep_row = subject_id in eligible_subject_ids

                if keep_row:
                    output_worksheet.append(list(row))
                    kept_rows += 1

            sheet_report.append(
                (input_worksheet.title, total_rows, kept_rows, filter_key)
            )

        append_dataframe_sheet(output_workbook, SELECTION_OUTPUT_SHEET, selection_output)
        append_dataframe_sheet(
            output_workbook,
            SUMMARY_OUTPUT_SHEET,
            pd.DataFrame(
                sheet_report,
                columns=["sheet_name", "rows_input", "rows_kept", "filter_key"],
            ),
        )
        output_workbook.save(output_path)
    finally:
        input_workbook.close()

    return sheet_report


def main() -> None:
    args = parse_args()

    input_path = args.input.expanduser().resolve()
    selection_path = args.selection.expanduser().resolve()
    key_path = args.key.expanduser().resolve()

    for path_label, path in (
        ("Input workbook", input_path),
        ("Selection workbook", selection_path),
        ("Key file", key_path),
    ):
        if not path.exists():
            raise SystemExit(f"{path_label} not found: {path}")

    output_path = (
        args.output.expanduser().resolve()
        if args.output is not None
        else default_output_path(input_path, args.min_age, args.max_age)
    )

    selection_frame, selection_sheet_name = read_selection_frame(
        selection_path, args.selection_sheet
    )
    key_frame = read_key_frame(key_path)
    selection_with_mapping, filtered_selection, patient_id_width = prepare_selection(
        selection_frame, key_frame, args.min_age, args.max_age
    )

    if filtered_selection.empty:
        raise SystemExit(
            "No patients remain after applying the key mapping and age filter."
        )

    eligible_study_ids = set(
        filtered_selection[STUDY_ID_NEW_COLUMN].dropna().astype(str).tolist()
    )

    eligible_subject_ids = build_eligible_subject_ids(input_path, eligible_study_ids)

    selection_output = filtered_selection.copy()
    if patient_id_width > 0:
        selection_output[AZGNR_COLUMN] = normalize_series(
            selection_output[AZGNR_COLUMN], width=patient_id_width
        )
    selection_output = selection_output.drop(
        columns=["patient_identifier_normalized"], errors="ignore"
    )

    sheet_report = stream_filter_workbook(
        input_path,
        output_path,
        eligible_study_ids,
        eligible_subject_ids,
        selection_output,
    )

    print(f"Input workbook:             {input_path}")
    print(f"Selection workbook:         {selection_path} [{selection_sheet_name}]")
    print(f"Key file:                   {key_path}")
    print(f"Output workbook:            {output_path}")
    print(f"Selection rows before age:  {len(selection_with_mapping)}")
    print(
        "Selection rows with key:    "
        f"{int(selection_with_mapping['has_study_id_new'].sum())}"
    )
    print(f"Selection rows kept:        {len(filtered_selection)}")
    print(f"Eligible study ids:         {len(eligible_study_ids)}")
    print(f"Eligible subject ids:       {len(eligible_subject_ids)}")
    print("")
    print("Rows kept per sheet:")
    for sheet_name, total_rows, kept_rows, filter_key in sheet_report:
        print(f"- {sheet_name}: {kept_rows}/{total_rows} rows via {filter_key}")


if __name__ == "__main__":
    main()
